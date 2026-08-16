"""
Parser for manuelt kopiert FINN-tekst -> CSV/XLSX.

Dette scriptet gjor IKKE noe automatisk oppslag mot FINN. Det leser en
lokal tekstfil (data/finn_raw.txt) som du selv har limt inn tekst i
(f.eks. Ctrl+A / Ctrl+C fra en resultatside pa FINN), og strukturerer
den til CSV og Excel for videre analyse i f.eks. Power BI.

Strategien er a dele radataen i "annonseblokker" avgrenset av
"Legg til som favoritt." (som FINN skriver etter hver annonse),
i stedet for a gjette linje-for-linje bakover fra adressen. Det gjor
parseren mer robust mot at enkelte annonser mangler felter
(f.eks. ingen egen tittel, ingen soverom, ingen tomteareal).
"""

import csv
import re
from pathlib import Path
from datetime import date

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


RAW_FILE = Path("data/finn_raw.txt")
CSV_FILE = Path("data/finn_vega.csv")
XLSX_FILE = Path("data/finn_vega.xlsx")

# Stedsnavnet som brukes til a finne adresselinjer, f.eks. "X-veien 4, Vega".
# Bytt denne hvis du parser et annet sted enn Vega.
LOCATION = "Vega"

FIELDNAMES = [
    "FinnId",
    "Tittel",
    "Megler",
    "Pris",
    "Totalpris",
    "Adresse",
    "Boligtype",
    "Eierform",
    "Soverom",
    "Areal",
    "Tomteareal",
    "Status",
    "Kilde",
    "SistOppdatert",
    "Url",
    "Kommentar",
]

BROKER_KEYWORDS = [
    "Nordbohus",
    "EiendomsMegler",
    "DNB Eiendom",
    "Aktiv",
    "PrivatMegleren",
    "EIE",
    "Krogsveen",
    "Notar",
    "RE/MAX",
    "Meglerhuset",
    "Sem & Johnsen",
    "Nordvik",
]

# Linjer som alltid skal fjernes uansett hvor i blokken de dukker opp.
LINE_NOISE_PREFIXES = (
    "Bilde ",
    "Megler logo",
    "Visning etter avtale",
    "Legg til som favoritt",
    "Digital visning",
)

# "Statusmerker" pa annonsebildet, f.eks. "4 solgt - kun 2 igjen!".
# Disse skal som hovedregel IKKE tolkes som tittel, men brukes som fallback
# hvis annonsen mangler tydelig tittel.
BADGE_PATTERNS = [
    re.compile(r"\d+\s*solgt", re.IGNORECASE),
    re.compile(r"kun\s*\d+\s*igjen", re.IGNORECASE),
    re.compile(r"\d+\s*budrunde", re.IGNORECASE),
    re.compile(r"nedsatt\s*pris", re.IGNORECASE),
    re.compile(r"ny\s*pris", re.IGNORECASE),
]

LISTING_START_RE = re.compile(r"^Bilde\s+\d+\s+av\s+annonsen$")
LISTING_END_PREFIX = "Legg til som favoritt"


def clean_text(value):
    if value is None:
        return ""

    text = str(value)
    text = text.replace("\u202f", " ")
    text = text.replace("\xa0", " ")
    text = text.replace("–", "-")
    text = text.replace("∙", "|")
    text = re.sub(r"\s+", " ", text)

    return text.strip()


def parse_int(value):
    text = clean_text(value)
    digits = re.sub(r"[^\d]", "", text)

    if digits == "":
        return None

    return int(digits)


def is_badge(line):
    return any(pattern.search(line) for pattern in BADGE_PATTERNS)


def find_badge(block_lines):
    for line in block_lines:
        if is_badge(line):
            return line

    return None


def is_broker_name(line):
    line_lower = line.lower()

    for keyword in BROKER_KEYWORDS:
        if keyword.lower() in line_lower:
            return True

    return False


def looks_like_price_only(line):
    if re.fullmatch(r"[\d\s]+", line):
        return True

    if re.fullmatch(r"[\d\s]+\s*kr", line, flags=re.IGNORECASE):
        return True

    return False


# ---------------------------------------------------------------------
# Feltuttrekk fra en hel annonseblokk
# ---------------------------------------------------------------------

def extract_total_price(block_text):
    match = re.search(
        r"Totalpris:\s*([\d\s]+)",
        block_text,
        flags=re.IGNORECASE,
    )

    if match:
        return parse_int(match.group(1))

    return None


def extract_area(block_text):
    # Foretrekk boareal som star rett foran prisen, f.eks. "108 m²2 350 000 kr".
    # Dette er et tryggere signal, siden tomteareal normalt ikke star rett foran "kr".
    match = re.search(
        r"(\d{2,4})\s*m[²2]\s*[\d\s]{4,}\s*kr",
        block_text,
        flags=re.IGNORECASE,
    )

    if match:
        return parse_int(match.group(1))

    # Fjern "Tomt pa X m²"-uttrykk for det er tomteareal, ikke boareal.
    without_plot = re.sub(
        r"Tomt(?:a| på| p[aå])?\s*[\d\s]+\s*m[²2]",
        "",
        block_text,
        flags=re.IGNORECASE,
    )

    match = re.search(
        r"(\d{2,4})\s*m[²2]",
        without_plot,
        flags=re.IGNORECASE,
    )

    if match:
        return parse_int(match.group(1))

    return None


def extract_plot_area(block_text):
    match = re.search(
        r"Tomt(?:a| på| p[aå])?\s*([\d\s]+)\s*m[²2]",
        block_text,
        flags=re.IGNORECASE,
    )

    if match:
        return parse_int(match.group(1))

    return None


def extract_bedrooms(block_text):
    match = re.search(
        r"(\d+)\s*soverom",
        block_text,
        flags=re.IGNORECASE,
    )

    if match:
        return parse_int(match.group(1))

    return None


def extract_eierform(block_text):
    known_values = [
        "Selveier",
        "Andel",
        "Aksje",
        "Obligasjon",
        "Annet",
    ]

    for value in known_values:
        if re.search(
            rf"\b{re.escape(value)}\b",
            block_text,
            flags=re.IGNORECASE,
        ):
            return value

    return None


def extract_boligtype(block_text):
    known_values = [
        "Tomannsbolig",
        "Gårdsbruk/Småbruk",
        "Gardsbruk/Smabruk",
        "Enebolig",
        "Leilighet",
        "Rekkehus",
        "Fritidsbolig",
        "Hytte",
        "Tomt",
        "Garasje/Parkering",
        "Produksjon/Industri",
    ]

    for value in known_values:
        if re.search(
            re.escape(value),
            block_text,
            flags=re.IGNORECASE,
        ):
            if value == "Gardsbruk/Smabruk":
                return "Gårdsbruk/Småbruk"

            return value

    return None


def extract_price(block_text):
    total_match = re.search(
        r"Totalpris:",
        block_text,
        flags=re.IGNORECASE,
    )

    before_totalpris = block_text[: total_match.start()] if total_match else block_text

    # Dekker tilfeller uten mellomrom mellom "m²" og prisen, f.eks. "122 m²2 850 000 kr"
    area_price_match = re.search(
        r"\d{2,4}\s*m[²2]\s*([\d\s]{5,})\s*kr",
        before_totalpris,
        flags=re.IGNORECASE,
    )

    if area_price_match:
        return parse_int(area_price_match.group(1))

    price_candidates_with_kr = re.findall(
        r"(\d[\d\s]{4,})\s*kr",
        before_totalpris,
        flags=re.IGNORECASE,
    )

    if price_candidates_with_kr:
        return parse_int(price_candidates_with_kr[-1])

    standalone_candidates = re.findall(
        r"(?m)^\s*(\d[\d\s]{5,})\s*$",
        before_totalpris,
    )

    if standalone_candidates:
        return parse_int(standalone_candidates[-1])

    return None


# ---------------------------------------------------------------------
# Blokk-splitting og parsing
# ---------------------------------------------------------------------

def split_into_blocks(raw_text):
    """Deler radataen i en liste av annonseblokker."""

    lines = [clean_text(line) for line in raw_text.splitlines()]
    lines = [line for line in lines if line != ""]

    start_idx = None

    for index, line in enumerate(lines):
        if LISTING_START_RE.match(line):
            start_idx = index
            break

    if start_idx is None:
        return []

    lines = lines[start_idx:]

    blocks = []
    current = []

    for line in lines:
        current.append(line)

        if line.startswith(LISTING_END_PREFIX):
            blocks.append(current)
            current = []

    # Ufullstendig siste blokk ignoreres bevisst.
    return blocks


def find_address(lines):
    address_re = re.compile(
        rf",\s*{re.escape(LOCATION)}\b",
        re.IGNORECASE,
    )

    for index, line in enumerate(lines):
        if address_re.search(line):
            return line, index

    return None, None


def find_broker(block_lines):
    for index, line in enumerate(block_lines):
        if line == "Megler logo" and index + 1 < len(block_lines):
            candidate = block_lines[index + 1]

            if candidate and not is_badge(candidate):
                return candidate

    for line in block_lines:
        if is_broker_name(line):
            return line

    return None


def parse_block(block_lines, row_number):
    broker = find_broker(block_lines)
    badge = find_badge(block_lines)

    content_lines = [
        line
        for line in block_lines
        if not line.startswith(LINE_NOISE_PREFIXES)
        and line != broker
        and not is_badge(line)
    ]

    address, address_pos = find_address(content_lines)

    title = None

    if address_pos is not None:
        for line in content_lines[:address_pos]:
            if looks_like_price_only(line):
                continue

            if len(line) < 6:
                continue

            title = line
            break

    # Fallback: Hvis tittel mangler, bruk statusmerke/badge som tittel.
    # Eksempel: "4 solgt - kun 2 igjen!"
    used_badge_as_title = False

    if title is None and badge is not None:
        title = badge
        used_badge_as_title = True

    if used_badge_as_title:
        comment = "Tittel hentet fra statusmerke - kontroller manuelt"
    elif not address:
        comment = "Mangler adresse - sjekk manuelt"
    else:
        comment = ""

    block_text = "\n".join(block_lines)

    return {
        "FinnId": f"ANNONSE_{row_number:03d}",
        "Tittel": title,
        "Megler": broker,
        "Pris": extract_price(block_text),
        "Totalpris": extract_total_price(block_text),
        "Adresse": address,
        "Boligtype": extract_boligtype(block_text),
        "Eierform": extract_eierform(block_text),
        "Soverom": extract_bedrooms(block_text),
        "Areal": extract_area(block_text),
        "Tomteareal": extract_plot_area(block_text),
        "Status": "Til salgs",
        "Kilde": "FINN manuelt kopiert tekst",
        "SistOppdatert": date.today().isoformat(),
        "Url": "",
        "Kommentar": comment,
    }


def parse_rows(raw_text):
    blocks = split_into_blocks(raw_text)

    rows = []

    for row_number, block_lines in enumerate(blocks, start=1):
        rows.append(parse_block(block_lines, row_number))

    return rows


# ---------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------

def write_csv(rows):
    CSV_FILE.parent.mkdir(parents=True, exist_ok=True)

    with CSV_FILE.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows):
    XLSX_FILE.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "finn_vega"

    ws.append(FIELDNAMES)

    for row in rows:
        ws.append([row.get(field) for field in FIELDNAMES])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 16,  # FinnId
        "B": 70,  # Tittel
        "C": 30,  # Megler
        "D": 14,  # Pris
        "E": 14,  # Totalpris
        "F": 30,  # Adresse
        "G": 22,  # Boligtype
        "H": 14,  # Eierform
        "I": 10,  # Soverom
        "J": 10,  # Areal
        "K": 14,  # Tomteareal
        "L": 14,  # Status
        "M": 28,  # Kilde
        "N": 16,  # SistOppdatert
        "O": 45,  # Url
        "P": 40,  # Kommentar
    }

    for column_letter, width in widths.items():
        ws.column_dimensions[column_letter].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_num in range(2, ws.max_row + 1):
        ws[f"D{row_num}"].number_format = "#,##0"
        ws[f"E{row_num}"].number_format = "#,##0"
        ws[f"I{row_num}"].number_format = "0"
        ws[f"J{row_num}"].number_format = "0"
        ws[f"K{row_num}"].number_format = "#,##0"

    if ws.max_row >= 2:
        table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

        table = Table(displayName="tblFinnVega", ref=table_ref)

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        table.tableStyleInfo = style
        ws.add_table(table)

    ws.freeze_panes = "A2"

    readme = wb.create_sheet("README")
    readme["A1"] = "FINN Vega parser"
    readme["A1"].font = Font(size=16, bold=True, color="1F4E78")

    readme["A3"] = "Bruk"
    readme["B3"] = (
        "Lim inn kopiert tekst fra FINN i data/finn_raw.txt. "
        "GitHub Action genererer CSV og Excel."
    )

    readme["A5"] = "Viktig"
    readme["B5"] = (
        "Dette er parsing av manuelt kopiert tekst, ikke automatisk scraping mot FINN."
    )

    readme["A7"] = "Power BI"
    readme["B7"] = "Koble Power BI mot data/finn_vega.xlsx eller data/finn_vega.csv."

    readme["A9"] = "Feilsøking"
    readme["B9"] = (
        "Hvis en rad mangler adresse eller tittel, sjekk kolonnen 'Kommentar' "
        "og se om annonsen manglet et 'Legg til som favoritt.'-avslutning i radataen."
    )

    readme.column_dimensions["A"].width = 18
    readme.column_dimensions["B"].width = 100

    for row in readme.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(XLSX_FILE)


def main():
    if not RAW_FILE.exists():
        raise FileNotFoundError(
            f"Fant ikke {RAW_FILE}. Opprett filen og lim inn kopiert tekst fra FINN."
        )

    raw_text = RAW_FILE.read_text(encoding="utf-8")
    rows = parse_rows(raw_text)

    write_csv(rows)
    write_xlsx(rows)

    print(f"Fant {len(rows)} annonser")
    print(f"Skrev {CSV_FILE}")
    print(f"Skrev {XLSX_FILE}")


if __name__ == "__main__":
    main()
